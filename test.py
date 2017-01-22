def main():
    import general
    import re

    # square root calculation
    general.square_root(4, 3)

    # regex examples
    res1 = re.compile(r'(\d\d\d)-(\d\d\d-\d\d\d\d)')
    mo1 = res1.search('My number is 415-555-4242 or 212-555-0000.')
    print(mo1.group())
    area, number = mo1.groups()
    print(mo1.group(1))
    print(area)
    print(mo1.group(2))
    print(number)
    mo1_1 = res1.findall('My number is 415-555-4242 or 212-555-0000.')
    print(mo1_1)

    res2 = re.compile(r'Bat(man|mobile|copter|arang)')
    mo2 = res2.search('Batmobile lost his wheel')
    print(mo2.group())
    print(mo2.group(1))

    res3 = re.compile(r'Bat(wo)?man')
    mo3 = res3.search('The Adventures of Batman')
    print(mo3.group())

    res4 = re.compile(r'\d+\s\w+')
    mo4 = res4.findall('12 drummers, 11 pipers, 10 lords, 9 ladies, 8 maids, 7 swans, 6 geese, 5 rings, 4 birds, 3 hens, 2 doves, 1 partridge')
    print(mo4)

    # exclude vowels
    res5 = re.compile(r'[^aeiouAEIOU]')
    mo5 = res5.findall('Quick brown fox jumps over the lazy dog')
    print(mo5)

    # begins and ends with a numerical
    res6 = re.compile(r'^\d+$')
    print(res6.search('12vb34') is None)

    # using thw wildcard
    res7 = re.compile(r'.at')
    mo7 = res7.findall('The cat in the hat sat on the flat mat.')
    print(mo7)

    # greedy vs non-greedy evaluation & catch-all combo (.*)
    res8_1 = re.compile(r'<.*?>')
    mo8 = res8_1.search('<To serve> and protect.>')
    print(mo8.group())
    res8_2 = re.compile(r'<.*>')
    mo8 = res8_2.search('<To serve> and protect.>')
    print(mo8.group())

    # replace all occurrences with a string
    res9_1 = re.compile(r'Agent \w+')
    mo9 = res9_1.sub('CENSORED', 'Agent Alice gave her secret key to Agent Bob')
    print(mo9)

    res9_2 = re.compile(r'Agent (\w)\w*')
    mo9 = res9_2.sub(r'\1****', 'Agent Alice gave her secret key to Agent Bob')
    print(mo9)

    # regex string over multiple lines
    # phone number in the format (555) 555-5555 x55
    res10_1 = re.compile(r'''(
        (\d{3}|\(\d{3}\))?              # area code
        (\s|-|\.)?                      # separator
        \d{3}                           # first three digits
        (\s|-|\.)                       # separator
        \d{4}                           # last four digits
        (\s*(ext|x|ext.)\s*\d{2,5})?    # extension
    )''', re.VERBOSE)
    # email address
    res10_2 = re.compile(r'''(
        [a-zA-Z0-9._%+-]+   # user
        @                   # at
        [a-zA-Z0-9.-]+      # domain
        (\.[a-zA-Z]{2,4})   # dot
    )''', re.VERBOSE)

    # parse xlsx
    import openpyxl
    from openpyxl.cell import get_column_letter, column_index_from_string
    wb = openpyxl.load_workbook('test.xlsx')
    print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Sheet1')
    print(sheet)
    print(sheet['A1'])
    print(sheet['A1'].value)
    c = sheet['B1']
    print(c.value)
    print('Row {}, Column {} is {}'.format(str(c.row), c.column, c.value))
    print('Cell {} is {}'.format(c.coordinate, c.value))
    print(sheet['C1'].value)
    print(get_column_letter(sheet.max_column))

    general.parse_excel('test.xlsx')
    general.list_excel_rows('test.xlsx')
    general.list_excel_columns('test.xlsx')

if __name__ == '__main__':
    main()
